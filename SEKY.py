from Levenshtein import distance
from openpyxl import Workbook
import fitz
from pytesseract import image_to_string
from re import search, split, IGNORECASE
from os import listdir


def main():
    while True:
        try:
            dir_path = input("Write the path to your PDFs directory: ")
            pdf_ls= get_pdf_ls(dir_path)

            kw_input = ""
            while not kw_input:
                kw_input = input("Write your keywords separated with a space:")

            key_words = get_keywords(kw_input)
            output = set_output(dir_path, pdf_ls, key_words)

            return print(output)

        except (FileNotFoundError, OSError):
            continue


def set_output(dir_path, pdf_ls, key_words):

    if len(pdf_ls) > 0:

        output, cnt_output, excel_wb = get_output(pdf_ls, key_words, dir_path)

        if cnt_output == 0:
            output = "There is no match, try with others keywords"
        else:
            #excel output
            xls_name = "_".join(key_words)
            excel_path = dir_path + "/keywords_" + xls_name + ".xls"
            excel_wb.save(excel_path)

            #print output
            notif = "\nYour results has been saved in the selected directory in an Excel file\n"
            output = f"\n{cnt_output} PDFs match the keywords:\n\n" + output + notif
    else:
        output = "The selected directory doesn't have any PDF file"

    return output


def get_output(pdf_ls, key_words, dir_path):

    #Getting a excel workbook and sheet
    excel_wb = Workbook()
    sheet = excel_wb.active
    #Writing the first row
    sheet.cell(row= 1, column=1).value = "Hiperlink (click to open the PDF)"
    sheet.cell(row= 1, column=2).value = "Keywords (the keywords that you searched)"
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 100

    #Getting and writing the output in "xls" file row by row
    cnt_output = 0
    output = ""
    rw = 1
    for pdf in pdf_ls:

        str_ls = get_str_ls(pdf, dir_path)
        matches_ls = get_matches_ls(str_ls, key_words)

        if len(matches_ls)/len(key_words) < 1:
            continue
        else:
            output += f"{pdf}\n"
            cnt_output += 1
            rw += 1
            sheet.cell(row= rw, column=1).hyperlink = dir_path + "/" + pdf
            sheet.cell(row= rw, column=1).value = pdf
            sheet.cell(row= rw, column=1).style = "Hyperlink"
            sheet.cell(row= rw, column=2).value = '-'.join(matches_ls)

    return output, cnt_output, excel_wb


def get_matches_ls(str_ls, key_words):

    matches_ls = []
    for s in str_ls:
        for k in key_words:
            d = distance(k,s)
            m = d/len(k)
            # treshold
            if m < 0.4:
                if k in matches_ls:
                    continue
                else:
                    matches_ls.append(k)
    return matches_ls


def get_str_ls(pdf, dir_path):
    #open a document and return the first page
    pdf_path = dir_path + "/" + pdf
    doc = fitz.open(pdf_path)
    page_1 = doc[0]
    str_val = page_1.get_text("text")

    #Render no searchable PDF to an image and use OCR to get the text
    if len(str_val) == 0:
        pix = page_1.get_pixmap()
        image_path = dir_path + "/page.png"
        pix.save(image_path)
        str_val = image_to_string(image_path, lang= "spa")

    str_ls = split(r"\W+", str_val.casefold())

    return str_ls


def get_pdf_ls(path_dir):

    dir_ls = listdir(path_dir)
    pdf_ls = []
    for file in dir_ls:
        if search(r"\w\.pdf$", file, IGNORECASE):
            pdf_ls.append(file)

    return pdf_ls


def get_keywords(kw_input):

    key_words = split(r"\W+", kw_input.casefold().strip())

    return key_words





if __name__ == "__main__":
    main()




